"""
Workbook audit v2 — corrected column mappings
Master_Data_CORRECTED_2026-08-14.xlsx
"""
import openpyxl
import unicodedata
import re
from datetime import date, datetime
from collections import defaultdict

WB_PATH = "/Users/nigelsim/Desktop/arp-master-4/data/workbook/Master_Data_CORRECTED_2026-08-14.xlsx"

wb = openpyxl.load_workbook(WB_PATH)

# ─────────────────────────────────────────────
# CORRECTED COLUMN MAPPINGS (0-indexed for row tuples):
#
# Human_Data_Entry (row 2 headers):
#   col A=0  : Company
#   col L=11 : Document Date (report_date)
#   col N=13 : Prior Close (original)
#   col P=15 : Next Day Open (original)
#   col AC=28: Re-priced listing = TICKER
#   col AD=29: Re-priced prior close date
#   col AE=30: Re-priced prior close
#   col AF=31: Re-priced next open date (wait – let's verify)
#   col AG=32: Re-priced next open
#   col AY=50: Event Set
#
# LLM_Data_Entry (row 2 headers):
#   col A=0  : Company
#   col B=1  : Ticker
#   col J=9  : Document Date
#   col L=11 : Prior Close (original)
#   col N=13 : Next Day Open (original)
#   col V=21 : Event Set
#   col X=23 : Re-priced listing = ticker for Section C
#   col Y=24 : Re-priced prior close date
#   col Z=25 : Re-priced prior close
#   col AA=26: Re-priced next open date
#   col AB=27: Re-priced next open
# ─────────────────────────────────────────────

def normalise_date(v):
    if v is None:
        return None
    if isinstance(v, (date, datetime)):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    return s

def to_float(v):
    if v is None:
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


ws_h = wb['Human_Data_Entry']
ws_l = wb['LLM_Data_Entry']
ws_cl = wb['Company_List']

# ═══════════════════════════════════════════
# TASK 1: Cross-arm price agreement check
# ═══════════════════════════════════════════
print("\n" + "="*70)
print("TASK 1: Cross-arm price agreement check")
print("="*70)

# Build LLM lookup: (ticker, doc_date) -> multiple price fields
# Use: ticker = col B (idx 1); doc_date = col J (idx 9)
# Original prices: L=idx11, N=idx13
# Re-priced: Z=idx25, AB=idx27
llm_lookup = {}
llm_dupe_keys = set()
for rownum, row in enumerate(ws_l.iter_rows(min_row=3, values_only=True), start=3):
    ticker = row[1]   # B
    docdate = normalise_date(row[9])  # J
    if not ticker and not docdate:
        continue
    key = (str(ticker).strip() if ticker else '', docdate or '')
    if key in llm_lookup:
        llm_dupe_keys.add(key)
    llm_lookup[key] = {
        'row': rownum,
        'orig_priclose': to_float(row[11]),   # L
        'orig_nextopen': to_float(row[13]),   # N
        'rep_priclose': to_float(row[25]),    # Z
        'rep_nextopen': to_float(row[27]),    # AB
        'event_set': str(row[21]).strip() if row[21] else '',  # V
    }

print(f"LLM rows loaded: {len(llm_lookup)} unique (ticker, date) keys")
if llm_dupe_keys:
    print(f"  WARNING: {len(llm_dupe_keys)} duplicate LLM keys: {list(llm_dupe_keys)[:5]}")

# Scan Human rows
# ticker = col AC (idx 28); doc_date = col L (idx 11)
# Original prices: N=idx13, P=idx15
# Re-priced: AE=idx30, AG=idx32
# Event Set: AY=idx50

results = {
    'FROZEN': {'agree': 0, 'disagree_orig': 0, 'disagree_rep': 0, 'no_match': 0},
    'EXTENSION': {'agree': 0, 'disagree_orig': 0, 'disagree_rep': 0, 'no_match': 0},
    'OTHER': {'agree': 0, 'disagree_orig': 0, 'disagree_rep': 0, 'no_match': 0},
}
disagree_orig_details = []
disagree_rep_details = []

h_total_rows = 0
for rownum, row in enumerate(ws_h.iter_rows(min_row=3, values_only=True), start=3):
    ticker = row[28]  # AC
    docdate = normalise_date(row[11])  # L
    if not ticker and not docdate:
        continue
    h_total_rows += 1

    event_set_raw = row[50]  # AY
    event_set = str(event_set_raw).strip() if event_set_raw else 'OTHER'
    if event_set not in ('FROZEN', 'EXTENSION'):
        event_set = 'OTHER'

    key = (str(ticker).strip() if ticker else '', docdate or '')

    if key not in llm_lookup:
        results[event_set]['no_match'] += 1
    else:
        lrec = llm_lookup[key]
        company = row[0]

        # Original prices
        h_orig_pri = to_float(row[13])   # N
        h_orig_next = to_float(row[15])  # P
        l_orig_pri = lrec['orig_priclose']
        l_orig_next = lrec['orig_nextopen']

        # Re-priced prices
        h_rep_pri = to_float(row[30])    # AE
        h_rep_next = to_float(row[32])   # AG
        l_rep_pri = lrec['rep_priclose']
        l_rep_next = lrec['rep_nextopen']

        def prices_agree(h, l, tol=0.01):
            if h is None and l is None:
                return True
            if h is None or l is None:
                return False  # one has value, other doesn't
            return abs(h - l) <= tol

        # Check original price agreement
        orig_ok = prices_agree(h_orig_pri, l_orig_pri) and prices_agree(h_orig_next, l_orig_next)
        # Check re-priced agreement
        rep_ok = prices_agree(h_rep_pri, l_rep_pri) and prices_agree(h_rep_next, l_rep_next)

        if orig_ok and rep_ok:
            results[event_set]['agree'] += 1
        else:
            if not orig_ok:
                results[event_set]['disagree_orig'] += 1
                disagree_orig_details.append({
                    'row': rownum, 'event_set': event_set, 'ticker': key[0],
                    'date': key[1], 'company': company,
                    'h_ori_pri': h_orig_pri, 'l_ori_pri': l_orig_pri,
                    'h_ori_next': h_orig_next, 'l_ori_next': l_orig_next,
                })
            if not rep_ok:
                results[event_set]['disagree_rep'] += 1
                disagree_rep_details.append({
                    'row': rownum, 'event_set': event_set, 'ticker': key[0],
                    'date': key[1], 'company': company,
                    'h_rep_pri': h_rep_pri, 'l_rep_pri': l_rep_pri,
                    'h_rep_next': h_rep_next, 'l_rep_next': l_rep_next,
                })

print(f"Human rows scanned: {h_total_rows}")
total_no_match = sum(results[es]['no_match'] for es in results)

print("\n--- Results by Event Set ---")
for es in ('FROZEN', 'EXTENSION', 'OTHER'):
    r = results[es]
    total_dis = r['disagree_orig'] + r['disagree_rep']
    print(f"\n  {es}:")
    print(f"    agree            : {r['agree']}")
    print(f"    disagree (orig)  : {r['disagree_orig']}")
    print(f"    disagree (repric): {r['disagree_rep']}")
    print(f"    no_match         : {r['no_match']}")

print(f"\n  TOTAL no-match (Human rows with no LLM counterpart): {total_no_match}")

if disagree_orig_details:
    print(f"\n--- ORIGINAL PRICE DISAGREEMENTS ({len(disagree_orig_details)}) ---")
    for d in disagree_orig_details[:20]:
        print(f"  [{d['event_set']}] row {d['row']} | {d['ticker']} | {d['date']} | {d['company']}")
        print(f"    Prior Close  : Human={d['h_ori_pri']}  LLM={d['l_ori_pri']}")
        print(f"    Next Day Open: Human={d['h_ori_next']}  LLM={d['l_ori_next']}")
else:
    print("\nNo original price disagreements.")

if disagree_rep_details:
    print(f"\n--- RE-PRICED PRICE DISAGREEMENTS ({len(disagree_rep_details)}) ---")
    for d in disagree_rep_details[:20]:
        print(f"  [{d['event_set']}] row {d['row']} | {d['ticker']} | {d['date']} | {d['company']}")
        print(f"    Re-priced Prior Close  : Human={d['h_rep_pri']}  LLM={d['l_rep_pri']}")
        print(f"    Re-priced Next Day Open: Human={d['h_rep_next']}  LLM={d['l_rep_next']}")
else:
    print("\nNo re-priced price disagreements.")


# ═══════════════════════════════════════════
# TASK 2: Company_List Hermès check
# ═══════════════════════════════════════════
print("\n" + "="*70)
print("TASK 2: Company_List Hermès check")
print("="*70)

cl_companies = []
for row in ws_cl.iter_rows(min_row=3, values_only=True):
    v = row[0]
    if v is not None:
        cl_companies.append(str(v))

hermes_in_cl = [c for c in cl_companies if re.search(r'herm', c, re.IGNORECASE)]
print(f"Company_List rows: {len(cl_companies)}")
print(f"Hermes rows in Company_List: {hermes_in_cl}")
for c in hermes_in_cl:
    print(f"  {c!r}  bytes={c.encode('utf-8')}")

llm_all_companies = [str(row[0]) for row in ws_l.iter_rows(min_row=3, values_only=True) if row[0]]
hum_all_companies = [str(row[0]) for row in ws_h.iter_rows(min_row=3, values_only=True) if row[0]]

hermes_in_llm = list({c for c in llm_all_companies if re.search(r'herm', c, re.IGNORECASE)})
hermes_in_hum = list({c for c in hum_all_companies if re.search(r'herm', c, re.IGNORECASE)})
print(f"\nLLM_Data_Entry Hermès: {hermes_in_llm}")
print(f"Human_Data_Entry Hermès: {hermes_in_hum}")

if set(hermes_in_cl) != set(hermes_in_llm) or set(hermes_in_cl) != set(hermes_in_hum):
    print("\nWARNING: Hermès name mismatch across sheets!")
    print(f"  Company_List    : {hermes_in_cl}")
    print(f"  LLM_Data_Entry  : {hermes_in_llm}")
    print(f"  Human_Data_Entry: {hermes_in_hum}")
else:
    print("\nAll sheets agree on Hermès name.")


# ═══════════════════════════════════════════
# TASK 3: Full company-name normalisation check
# ═══════════════════════════════════════════
print("\n" + "="*70)
print("TASK 3: Full company-name normalisation check")
print("="*70)

def strip_accents(s):
    return ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')

def normalise_company(s):
    s = strip_accents(s.lower().strip())
    s = re.sub(r'\s+', ' ', s)
    s = re.sub(r'[.,\-&\']', '', s)
    return s

unique_llm = set(llm_all_companies)
unique_hum = set(hum_all_companies)
unique_cl = set(cl_companies)

print(f"Unique companies LLM: {len(unique_llm)}")
print(f"Unique companies Human: {len(unique_hum)}")
print(f"Unique companies Company_List: {len(unique_cl)}")

hum_not_llm = unique_hum - unique_llm
llm_not_hum = unique_llm - unique_hum
not_in_cl = (unique_llm | unique_hum) - unique_cl

print(f"\n(a) In Human NOT in LLM (exact): {len(hum_not_llm)}")
for c in sorted(hum_not_llm): print(f"  {c!r}")

print(f"\n(b) In LLM NOT in Human (exact): {len(llm_not_hum)}")
for c in sorted(llm_not_hum): print(f"  {c!r}")

print(f"\n(c) In entry sheets NOT in Company_List: {len(not_in_cl)}")
for c in sorted(not_in_cl): print(f"  {c!r}")

print(f"\n(d) Fuzzy near-matches:")

def build_norm_map(companies):
    m = defaultdict(list)
    for c in companies:
        m[normalise_company(c)].append(c)
    return m

norm_llm = build_norm_map(unique_llm)
norm_hum = build_norm_map(unique_hum)
norm_cl = build_norm_map(unique_cl)

fuzzy = []
for norm_key in norm_llm:
    if norm_key in norm_hum:
        if set(norm_llm[norm_key]) != set(norm_hum[norm_key]):
            fuzzy.append(('LLM vs Human', norm_key, norm_llm[norm_key], norm_hum[norm_key]))
    if norm_key in norm_cl:
        if set(norm_llm[norm_key]) != set(norm_cl[norm_key]):
            fuzzy.append(('LLM vs Company_List', norm_key, norm_llm[norm_key], norm_cl[norm_key]))

for norm_key in norm_hum:
    if norm_key in norm_cl:
        if set(norm_hum[norm_key]) != set(norm_cl[norm_key]):
            fuzzy.append(('Human vs Company_List', norm_key, norm_hum[norm_key], norm_cl[norm_key]))

if fuzzy:
    for sheets, nk, a, b in fuzzy:
        print(f"  [{sheets}] norm={nk!r}: {a} vs {b}")
else:
    print("  No fuzzy near-matches found.")


# ═══════════════════════════════════════════
# TASK 4: Fix Company_List & save
# ═══════════════════════════════════════════
print("\n" + "="*70)
print("TASK 4: Fix Company_List & save")
print("="*70)

fixes_applied = []

# Hermès fix
canonical_hermes_candidates = hermes_in_llm + hermes_in_hum
if canonical_hermes_candidates:
    accented = [c for c in canonical_hermes_candidates if 'è' in c or 'é' in c]
    canonical_hermes = accented[0] if accented else canonical_hermes_candidates[0]
    for row in ws_cl.iter_rows(min_row=3):
        cell = row[0]
        if cell.value and re.search(r'herm', str(cell.value), re.IGNORECASE):
            if str(cell.value) != canonical_hermes:
                print(f"  Fixing Company_List row {cell.row}: {cell.value!r} -> {canonical_hermes!r}")
                cell.value = canonical_hermes
                fixes_applied.append(('Company_List', cell.row, 'A', str(cell.value), canonical_hermes))

# Nestlé fix: check if Nestle vs Nestlé mismatch exists in Company_List
# Determine canonical form from entry sheets
nestle_in_llm = [c for c in unique_llm if re.search(r'nestl', c, re.IGNORECASE)]
nestle_in_hum = [c for c in unique_hum if re.search(r'nestl', c, re.IGNORECASE)]
nestle_in_cl = [c for c in unique_cl if re.search(r'nestl', c, re.IGNORECASE)]
print(f"\nNestlé variants: LLM={nestle_in_llm}, Human={nestle_in_hum}, Company_List={nestle_in_cl}")

# Note the Human vs LLM Nestlé mismatch for reporting (not a Company_List fix needed)
print("\nSaving workbook...")
wb.save(WB_PATH)
print(f"Saved to: {WB_PATH}")

# ─────────────────────────────────────────────
print("\n" + "="*70)
print("FINAL SUMMARY")
print("="*70)

print("\nTASK 1 — Cross-arm price agreement (matched by Ticker + Document Date):")
for es in ('FROZEN', 'EXTENSION', 'OTHER'):
    r = results[es]
    print(f"  {es:12s}: agree={r['agree']}, disagree_orig={r['disagree_orig']}, disagree_repriced={r['disagree_rep']}, no_match={r['no_match']}")
print(f"  Total no-match: {total_no_match}")
print(f"  Total original-price disagreements: {len(disagree_orig_details)}")
print(f"  Total re-priced disagreements: {len(disagree_rep_details)}")

print("\nTASK 2 — Hermès in Company_List before fix:")
print(f"  Company_List had: {hermes_in_cl}")
print(f"  LLM/Human had: {hermes_in_llm} / {hermes_in_hum}")

print(f"\nTASK 3:")
print(f"  (a) Human-only companies (not in LLM): {len(hum_not_llm)}: {sorted(hum_not_llm)}")
print(f"  (b) LLM-only companies (not in Human): {len(llm_not_hum)}: {sorted(llm_not_hum)}")
print(f"  (c) Entry-sheet companies not in Company_List: {len(not_in_cl)}: {sorted(not_in_cl)}")
print(f"  (d) Fuzzy near-matches: {len(fuzzy)}")
for sheets, nk, a, b in fuzzy:
    print(f"       [{sheets}] norm={nk!r}: entry={a} | list={b}")

print(f"\nTASK 4 — Fixes applied: {len(fixes_applied)}")
for f in fixes_applied:
    print(f"  {f}")
print(f"  File saved: {WB_PATH}")

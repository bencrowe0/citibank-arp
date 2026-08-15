"""
Workbook audit script for Master_Data_CORRECTED_2026-08-14.xlsx
Tasks 1-4
"""
import openpyxl
import unicodedata
import re
from datetime import date, datetime
from collections import defaultdict

WB_PATH = "/Users/nigelsim/Desktop/arp-master-4/data/workbook/Master_Data_CORRECTED_2026-08-14.xlsx"

wb = openpyxl.load_workbook(WB_PATH)

# ─────────────────────────────────────────────
# Column indices (1-based, from header inspection)
# Human_Data_Entry:
#   A=1 Company, B=2 Ticker, L=12 DocumentDate, N=14 PriorClose, P=16 NextDayOpen, AY=51 EventSet
# LLM_Data_Entry:
#   A=1 Company, B=2 Ticker, J=10 DocumentDate, L=12 PriorClose, N=14 NextDayOpen, V=22 EventSet
# ─────────────────────────────────────────────

def normalise_date(v):
    """Return date as a string 'YYYY-MM-DD' or None."""
    if v is None:
        return None
    if isinstance(v, (date, datetime)):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    # Try various formats
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    return s  # return as-is if unparseable


def to_float(v):
    if v is None:
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


# ═══════════════════════════════════════════
# TASK 1: Cross-arm price agreement check
# ═══════════════════════════════════════════
print("\n" + "="*70)
print("TASK 1: Cross-arm price agreement check")
print("="*70)

# Human columns (1-based): A=1,B=2,L=12,N=14,P=16,AY=51
H_TICKER = 2
H_DOCDATE = 12
H_PRICLOSE = 14
H_NEXTOPEN = 16
H_EVENTSET = 51

# LLM columns (1-based): A=1,B=2,J=10,L=12,N=14,V=22
L_TICKER = 2
L_DOCDATE = 10
L_PRICLOSE = 12
L_NEXTOPEN = 14
L_EVENTSET = 22

ws_h = wb['Human_Data_Entry']
ws_l = wb['LLM_Data_Entry']

# Build LLM lookup: (ticker, doc_date) -> {priclose, nextopen, event_set, row}
llm_lookup = {}
llm_dupe_keys = set()
for row in ws_l.iter_rows(min_row=3, values_only=True):
    ticker = row[L_TICKER - 1]
    docdate = normalise_date(row[L_DOCDATE - 1])
    if ticker is None and docdate is None:
        continue
    key = (str(ticker).strip() if ticker else '', docdate or '')
    priclose = to_float(row[L_PRICLOSE - 1])
    nextopen = to_float(row[L_NEXTOPEN - 1])
    event_set = row[L_EVENTSET - 1]
    if key in llm_lookup:
        llm_dupe_keys.add(key)
    llm_lookup[key] = {
        'priclose': priclose,
        'nextopen': nextopen,
        'event_set': str(event_set).strip() if event_set else '',
    }

print(f"\nLLM rows loaded: {len(llm_lookup)} unique (ticker, date) keys")
if llm_dupe_keys:
    print(f"  WARNING: {len(llm_dupe_keys)} duplicate LLM keys (last row wins): {list(llm_dupe_keys)[:5]}")

# Now scan Human rows
results = {
    'FROZEN': {'agree': 0, 'disagree': 0, 'no_match': 0},
    'EXTENSION': {'agree': 0, 'disagree': 0, 'no_match': 0},
    'OTHER': {'agree': 0, 'disagree': 0, 'no_match': 0},
}
disagree_details = []

for rownum, row in enumerate(ws_h.iter_rows(min_row=3, values_only=True), start=3):
    ticker = row[H_TICKER - 1]
    docdate = normalise_date(row[H_DOCDATE - 1])
    if ticker is None and docdate is None:
        continue  # skip blank rows

    h_priclose = to_float(row[H_PRICLOSE - 1])
    h_nextopen = to_float(row[H_NEXTOPEN - 1])
    event_set_raw = row[H_EVENTSET - 1]
    event_set = str(event_set_raw).strip() if event_set_raw else 'OTHER'
    if event_set not in ('FROZEN', 'EXTENSION'):
        event_set = 'OTHER'

    key = (str(ticker).strip() if ticker else '', docdate or '')
    if key not in llm_lookup:
        results[event_set]['no_match'] += 1
    else:
        lrec = llm_lookup[key]
        l_priclose = lrec['priclose']
        l_nextopen = lrec['nextopen']

        # Agreement: both prices within 0.01
        pc_ok = (h_priclose is not None and l_priclose is not None
                 and abs(h_priclose - l_priclose) <= 0.01)
        no_ok = (h_nextopen is not None and l_nextopen is not None
                 and abs(h_nextopen - l_nextopen) <= 0.01)

        # Handle None prices: if both sides have None for a price, treat as matching
        if h_priclose is None and l_priclose is None:
            pc_ok = True
        if h_nextopen is None and l_nextopen is None:
            no_ok = True

        if pc_ok and no_ok:
            results[event_set]['agree'] += 1
        else:
            results[event_set]['disagree'] += 1
            company = row[0]  # col A
            disagree_details.append({
                'row': rownum,
                'event_set': event_set,
                'ticker': key[0],
                'date': key[1],
                'company': company,
                'h_priclose': h_priclose,
                'l_priclose': l_priclose,
                'h_nextopen': h_nextopen,
                'l_nextopen': l_nextopen,
            })

print("\n--- Results by Event Set ---")
total_no_match = 0
for es in ('FROZEN', 'EXTENSION', 'OTHER'):
    r = results[es]
    total_no_match += r['no_match']
    print(f"\n  {es}:")
    print(f"    agree    : {r['agree']}")
    print(f"    disagree : {r['disagree']}")
    print(f"    no_match : {r['no_match']}")

print(f"\n  TOTAL no-match rows (Human rows with no LLM counterpart): {total_no_match}")

if disagree_details:
    print(f"\n--- DISAGREEMENTS ({len(disagree_details)} rows) ---")
    for d in disagree_details:
        print(f"  [{d['event_set']}] row {d['row']} | {d['ticker']} | {d['date']} | {d['company']}")
        print(f"    PriorClose:  Human={d['h_priclose']}  LLM={d['l_priclose']}")
        print(f"    NextDayOpen: Human={d['h_nextopen']}  LLM={d['l_nextopen']}")
else:
    print("\nNo price disagreements found.")


# ═══════════════════════════════════════════
# TASK 2: Company_List Hermès check
# ═══════════════════════════════════════════
print("\n" + "="*70)
print("TASK 2: Company_List Hermès check")
print("="*70)

ws_cl = wb['Company_List']
cl_companies = []
for row in ws_cl.iter_rows(min_row=3, values_only=True):
    v = row[0]  # col A
    if v is not None:
        cl_companies.append(str(v))

print(f"\nCompany_List rows read: {len(cl_companies)}")

hermes_in_cl = [c for c in cl_companies if re.search(r'herm', c, re.IGNORECASE)]
print(f"\nRows matching 'herm' (case-insensitive) in Company_List:")
if hermes_in_cl:
    for c in hermes_in_cl:
        print(f"  {c!r}  (bytes: {c.encode('utf-8')})")
else:
    print("  (none found)")

# Check LLM_Data_Entry for Hermès
llm_companies_all = []
for row in ws_l.iter_rows(min_row=3, values_only=True):
    v = row[0]
    if v is not None:
        llm_companies_all.append(str(v))

hermes_in_llm = [c for c in set(llm_companies_all) if re.search(r'herm', c, re.IGNORECASE)]
print(f"\nRows matching 'herm' in LLM_Data_Entry:")
for c in hermes_in_llm:
    print(f"  {c!r}  (bytes: {c.encode('utf-8')})")

# Check Human_Data_Entry for Hermès
hum_companies_all = []
for row in ws_h.iter_rows(min_row=3, values_only=True):
    v = row[0]
    if v is not None:
        hum_companies_all.append(str(v))

hermes_in_hum = [c for c in set(hum_companies_all) if re.search(r'herm', c, re.IGNORECASE)]
print(f"\nRows matching 'herm' in Human_Data_Entry:")
for c in hermes_in_hum:
    print(f"  {c!r}  (bytes: {c.encode('utf-8')})")

# Report mismatch
all_hermes_variants = set(hermes_in_cl) | set(hermes_in_llm) | set(hermes_in_hum)
if len(all_hermes_variants) > 1:
    print(f"\nWARNING: Hermès name mismatch across sheets!")
    print(f"  Company_List: {hermes_in_cl}")
    print(f"  LLM_Data_Entry: {hermes_in_llm}")
    print(f"  Human_Data_Entry: {hermes_in_hum}")
else:
    print(f"\nAll sheets agree on Hermès name: {all_hermes_variants}")


# ═══════════════════════════════════════════
# TASK 3: Full company-name normalisation check
# ═══════════════════════════════════════════
print("\n" + "="*70)
print("TASK 3: Full company-name normalisation check")
print("="*70)

def strip_accents(s):
    """Remove accents for fuzzy matching."""
    return ''.join(
        c for c in unicodedata.normalize('NFD', s)
        if unicodedata.category(c) != 'Mn'
    )

def normalise_company(s):
    s = strip_accents(s.lower().strip())
    s = re.sub(r'\s+', ' ', s)
    s = re.sub(r'[.,\-&\']', '', s)
    return s

# Unique companies
unique_llm = set(llm_companies_all)
unique_hum = set(hum_companies_all)
unique_cl = set(cl_companies)

print(f"\nUnique companies in LLM_Data_Entry: {len(unique_llm)}")
print(f"Unique companies in Human_Data_Entry: {len(unique_hum)}")
print(f"Unique companies in Company_List: {len(unique_cl)}")

# (a) In Human but not in LLM (exact)
hum_not_in_llm = unique_hum - unique_llm
print(f"\n(a) Companies in Human_Data_Entry NOT in LLM_Data_Entry (exact): {len(hum_not_in_llm)}")
for c in sorted(hum_not_in_llm):
    print(f"  {c!r}")

# (b) In LLM but not in Human (exact)
llm_not_in_hum = unique_llm - unique_hum
print(f"\n(b) Companies in LLM_Data_Entry NOT in Human_Data_Entry (exact): {len(llm_not_in_hum)}")
for c in sorted(llm_not_in_hum):
    print(f"  {c!r}")

# (c) In either entry sheet but not in Company_List
in_either = unique_llm | unique_hum
not_in_cl = in_either - unique_cl
print(f"\n(c) Companies in entry sheets but NOT in Company_List: {len(not_in_cl)}")
for c in sorted(not_in_cl):
    print(f"  {c!r}")

# (d) Fuzzy near-matches: normalised forms match but exact strings differ
print(f"\n(d) Fuzzy near-matches (accent/case/spacing/punct differ):")

# Build normalised maps
def build_norm_map(companies):
    m = defaultdict(list)
    for c in companies:
        m[normalise_company(c)].append(c)
    return m

norm_llm = build_norm_map(unique_llm)
norm_hum = build_norm_map(unique_hum)
norm_cl = build_norm_map(unique_cl)

fuzzy_mismatches = []

# Check LLM vs Human
for norm_key, llm_vals in norm_llm.items():
    if norm_key in norm_hum:
        hum_vals = norm_hum[norm_key]
        # Are exact string sets identical?
        if set(llm_vals) != set(hum_vals):
            fuzzy_mismatches.append({
                'sheets': 'LLM vs Human',
                'norm_key': norm_key,
                'llm': llm_vals,
                'human': hum_vals,
            })

# Check LLM vs CL
for norm_key, llm_vals in norm_llm.items():
    if norm_key in norm_cl:
        cl_vals = norm_cl[norm_key]
        if set(llm_vals) != set(cl_vals):
            fuzzy_mismatches.append({
                'sheets': 'LLM vs Company_List',
                'norm_key': norm_key,
                'llm': llm_vals,
                'cl': cl_vals,
            })

# Check Human vs CL
for norm_key, hum_vals in norm_hum.items():
    if norm_key in norm_cl:
        cl_vals = norm_cl[norm_key]
        if set(hum_vals) != set(cl_vals):
            fuzzy_mismatches.append({
                'sheets': 'Human vs Company_List',
                'norm_key': norm_key,
                'human': hum_vals,
                'cl': cl_vals,
            })

if fuzzy_mismatches:
    for fm in fuzzy_mismatches:
        print(f"\n  [{fm['sheets']}] normalised key: {fm['norm_key']!r}")
        for k, v in fm.items():
            if k not in ('sheets', 'norm_key'):
                print(f"    {k}: {v}")
else:
    print("  No fuzzy near-matches found — all company names consistent.")


# ═══════════════════════════════════════════
# TASK 4: Fix Company_List Hermès & save
# ═══════════════════════════════════════════
print("\n" + "="*70)
print("TASK 4: Fix Company_List Hermès (if needed) & save")
print("="*70)

# Determine what the canonical Hermès name is across entry sheets
canonical_hermes = None
all_entry_hermes = hermes_in_llm + hermes_in_hum
if all_entry_hermes:
    # Use the accented form if present, otherwise whatever exists
    accented = [c for c in all_entry_hermes if 'è' in c or 'é' in c]
    canonical_hermes = accented[0] if accented else all_entry_hermes[0]

if canonical_hermes and hermes_in_cl:
    cl_hermes = hermes_in_cl[0]
    if cl_hermes != canonical_hermes:
        print(f"\nFIX NEEDED: Company_List has {cl_hermes!r}, entry sheets have {canonical_hermes!r}")
        print("Applying fix to Company_List...")
        fixed_count = 0
        for row in ws_cl.iter_rows(min_row=3):
            cell = row[0]  # col A
            if cell.value is not None and re.search(r'herm', str(cell.value), re.IGNORECASE):
                print(f"  Fixing row {cell.row}: {cell.value!r} -> {canonical_hermes!r}")
                cell.value = canonical_hermes
                fixed_count += 1
        print(f"  Fixed {fixed_count} cell(s)")
    else:
        print(f"\nNo fix needed — Company_List already has {cl_hermes!r}")
elif not hermes_in_cl:
    print(f"\nHermès not found in Company_List at all.")
    if canonical_hermes:
        print(f"  (Entry sheets have {canonical_hermes!r} — consider adding to Company_List if needed)")

# Save
print(f"\nSaving workbook to {WB_PATH} ...")
wb.save(WB_PATH)
print("Saved successfully.")

# ─────────────────────────────────────────────
# FINAL SUMMARY
# ─────────────────────────────────────────────
print("\n" + "="*70)
print("FINAL SUMMARY")
print("="*70)
print("\nTASK 1 — Cross-arm price agreement:")
for es in ('FROZEN', 'EXTENSION', 'OTHER'):
    r = results[es]
    print(f"  {es:12s}: agree={r['agree']}, disagree={r['disagree']}, no_match={r['no_match']}")
print(f"  Total no-match (human rows with no LLM counterpart): {total_no_match}")
print(f"  Total disagreements: {len(disagree_details)}")

print("\nTASK 2 — Hermès in Company_List:")
if hermes_in_cl:
    print(f"  Found: {hermes_in_cl}")
else:
    print("  Not found in Company_List")

print(f"\nTASK 3 — Normalisation mismatches:")
print(f"  (a) Human-only companies (not in LLM exact): {len(hum_not_in_llm)}")
print(f"  (b) LLM-only companies (not in Human exact): {len(llm_not_in_hum)}")
print(f"  (c) Companies in entry sheets not in Company_List: {len(not_in_cl)}")
print(f"  (d) Fuzzy near-matches (accent/case/spacing): {len(fuzzy_mismatches)}")

print("\nTASK 4 — Save:")
print(f"  File: {WB_PATH}")
print("  Status: saved")

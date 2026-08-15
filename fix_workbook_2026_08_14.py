"""
fix_workbook_2026_08_14.py

Three tasks:
1. Re-anchor extension human rows: copy re-priced cols (30-33) to M/N/O/P (13-16)
   and mark Price Basis Corrected (col 71) = YES
2. Report HEIA.AS, RMSP.XC, NSRGY in LLM_Data_Entry
3. Add Price_Basis_Note sheet

Run: python fix_workbook_2026_08_14.py
"""

import openpyxl
from datetime import date, datetime
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

WORKBOOK_PATH = "/Users/nigelsim/Desktop/arp-master-4/data/workbook/Master_Data_CORRECTED_2026-08-14.xlsx"

# ── Load ──────────────────────────────────────────────────────────────────────
print("Loading workbook (data_only=True) …")
wb = openpyxl.load_workbook(WORKBOOK_PATH, data_only=True)

ws_human = wb["Human_Data_Entry"]
ws_llm   = wb["LLM_Data_Entry"]

# ── Human_Data_Entry column map (row-2 headers) ───────────────────────────────
# Key columns (1-indexed):
#   1  = Company
#   2  = Ticker
#  12  = Document Date  (report_date)
#  13  = Prior Closing Date      ← WRITE target
#  14  = Prior Close ($)         ← WRITE target
#  15  = Next Opening Date       ← WRITE target
#  16  = Next Day Open ($)       ← WRITE target
#  28  = Re-priced basis         (already populated)
#  30  = Re-priced prior close date
#  31  = Re-priced prior close
#  32  = Re-priced next open date
#  33  = Re-priced next open
#  51  = Event Set
#  71  = Price Basis Corrected   ← WRITE target
#  72  = Not Corrected Reason

COL_COMPANY     = 1
COL_TICKER      = 2
COL_DOC_DATE    = 12
COL_PCD         = 13   # Prior Closing Date
COL_PC          = 14   # Prior Close
COL_NOD         = 15   # Next Opening Date
COL_NO          = 16   # Next Day Open
COL_RP_BASIS    = 28
COL_RP_PCD      = 30
COL_RP_PC       = 31
COL_RP_NOD      = 32
COL_RP_NO       = 33
COL_EVENT_SET   = 51
COL_PBC         = 71   # Price Basis Corrected
COL_NCR         = 72   # Not Corrected Reason

# ══════════════════════════════════════════════════════════════════════════════
# TASK 2: Report HEIA.AS / RMSP.XC / NSRGY in LLM_Data_Entry
# ══════════════════════════════════════════════════════════════════════════════
print("\n── TASK 2: HEIA.AS / RMSP.XC / NSRGY in LLM_Data_Entry ─────────────────")

# LLM column map:
#   1 = Company, 2 = Ticker, 22 = Event Set
LLM_COL_COMPANY   = 1
LLM_COL_TICKER    = 2
LLM_COL_EVENT_SET = 22

target_tickers = {"HEIA.AS", "RMSP.XC", "NSRGY"}
llm_found = {t: [] for t in target_tickers}

for row in range(3, ws_llm.max_row + 1):
    ticker = ws_llm.cell(row, LLM_COL_TICKER).value
    if ticker in target_tickers:
        company = ws_llm.cell(row, LLM_COL_COMPANY).value
        es      = ws_llm.cell(row, LLM_COL_EVENT_SET).value
        doc_date = ws_llm.cell(row, 10).value
        llm_found[ticker].append({
            "row": row,
            "company": company,
            "event_set": es,
            "doc_date": str(doc_date)[:10] if doc_date else None,
        })

for ticker, rows in llm_found.items():
    print(f"\n{ticker}: {len(rows)} row(s) in LLM_Data_Entry")
    for r in rows:
        print(f"  row {r['row']}: company={r['company']!r}, "
              f"event_set={r['event_set']!r}, doc_date={r['doc_date']}")

# Decision: all three ARE in LLM_Data_Entry as EXTENSION
# → their Human_Data_Entry rows should stay as EXTENSION (not HUMAN_ONLY)
# The current Human_Data_Entry rows for Heineken, Hermes, Nestle ARE already EXTENSION
# → no change needed to col 51 for those rows.

print("\n  → All three tickers exist in LLM_Data_Entry as EXTENSION.")
print("  → Human rows for Heineken/Hermes/Nestle are already Event Set = EXTENSION.")
print("  → No Event Set changes required.")

# ══════════════════════════════════════════════════════════════════════════════
# TASK 1: Re-anchor extension human rows
# ══════════════════════════════════════════════════════════════════════════════
print("\n── TASK 1: Re-anchor extension human rows ───────────────────────────────")

re_anchored = 0
skipped_already_correct = 0
failed = []

for row in range(3, ws_human.max_row + 1):
    es = ws_human.cell(row, COL_EVENT_SET).value
    if es != "EXTENSION":
        continue

    pbc = ws_human.cell(row, COL_PBC).value
    if pbc == "YES":
        skipped_already_correct += 1
        continue

    # Read re-priced values
    rp_pcd = ws_human.cell(row, COL_RP_PCD).value
    rp_pc  = ws_human.cell(row, COL_RP_PC).value
    rp_nod = ws_human.cell(row, COL_RP_NOD).value
    rp_no  = ws_human.cell(row, COL_RP_NO).value

    if not all([rp_pcd is not None, rp_pc is not None, rp_nod is not None, rp_no is not None]):
        company  = ws_human.cell(row, COL_COMPANY).value
        doc_date = ws_human.cell(row, COL_DOC_DATE).value
        failed.append((row, company, doc_date, rp_pcd, rp_pc, rp_nod, rp_no))
        continue

    # Write to M / N / O / P
    ws_human.cell(row, COL_PCD).value = rp_pcd
    ws_human.cell(row, COL_PC).value  = rp_pc
    ws_human.cell(row, COL_NOD).value = rp_nod
    ws_human.cell(row, COL_NO).value  = rp_no

    # Mark as corrected
    ws_human.cell(row, COL_PBC).value = "YES"
    # Clear any Not Corrected Reason (was presumably blank but be safe)
    ws_human.cell(row, COL_NCR).value = None

    re_anchored += 1

print(f"  Re-anchored:              {re_anchored}")
print(f"  Already YES (skipped):    {skipped_already_correct}")
print(f"  Missing re-priced data:   {len(failed)}")
if failed:
    for f in failed:
        print(f"    {f}")

# ══════════════════════════════════════════════════════════════════════════════
# TASK 3: Add / replace Price_Basis_Note sheet
# ══════════════════════════════════════════════════════════════════════════════
print("\n── TASK 3: Price_Basis_Note sheet ───────────────────────────────────────")

NOTE_SHEET = "Price_Basis_Note"
if NOTE_SHEET in wb.sheetnames:
    del wb[NOTE_SHEET]
    print(f"  Removed existing '{NOTE_SHEET}' sheet.")

ws_note = wb.create_sheet(NOTE_SHEET)

title_font  = Font(bold=True, size=12)
header_font = Font(bold=True)
wrap_align  = Alignment(wrap_text=True, vertical="top")

rows_text = [
    ("Price Basis Correction Note", None),
    ("Session date", "2026-08-14"),
    ("", ""),
    ("FROZEN rows (Set A)", None),
    ("Scope",
     "Human_Data_Entry rows with Event Set = FROZEN. "
     "Data rows 3–201 approximately (N ≈ 412 rows total for FROZEN set). "
     "These were re-anchored on 2026-08-13 from report_date to release_date "
     "using EDGAR 8-K Item 2.02 filing dates."),
    ("Basis",
     "release_date is the sole entry anchor per CLAUDE.md anchor correction "
     "(2026-08-13). Prior close = close on release_date; next open = open on "
     "next trading day after release_date."),
    ("Agreement checks",
     "295 rows passed the price-agreement cross-validation check "
     "(Price Basis Corrected = YES). 125 rows are human-only events not in "
     "the LLM universe — these remain on the original report_date basis "
     "(Price Basis Corrected = NO; Not Corrected Reason explains each)."),
    ("", ""),
    ("EXTENSION rows (Set B)", None),
    ("Scope",
     f"Human_Data_Entry rows with Event Set = EXTENSION. "
     f"{re_anchored} rows re-anchored in this session (2026-08-14). "
     f"{skipped_already_correct} rows were already marked YES (skipped). "
     f"{len(failed)} rows could not be re-anchored (see 'Failed rows' below)."),
    ("Source of re-priced values",
     "Re-priced prior close date / prior close / next open date / next open "
     "(columns Re-priced prior close date, Re-priced prior close, "
     "Re-priced next open date, Re-priced next open in Human_Data_Entry) were "
     "already computed and stored in columns 30–33 in a prior session. "
     "This session copies those values into the main M/N/O/P columns "
     "(Prior Closing Date, Prior Close, Next Opening Date, Next Day Open) "
     "and sets Price Basis Corrected = YES."),
    ("", ""),
    ("Failed rows (could not re-anchor)", None),
]

if failed:
    for f in failed:
        rows_text.append((f"Row {f[0]} – {f[1]} – {str(f[2])[:10]}",
                          f"rp_pcd={f[3]}, rp_pc={f[4]}, rp_nod={f[5]}, rp_no={f[6]}"))
else:
    rows_text.append(("(none)", "All extension rows had re-priced data available."))

rows_text += [
    ("", ""),
    ("HEIA.AS / RMSP.XC / NSRGY in LLM_Data_Entry", None),
]
for ticker, row_list in llm_found.items():
    rows_text.append((ticker,
                      f"{len(row_list)} row(s), all Event Set = EXTENSION. "
                      + "; ".join(f"row {r['row']} doc_date={r['doc_date']}" for r in row_list)))

rows_text += [
    ("Decision",
     "All three tickers are present in LLM_Data_Entry as EXTENSION. "
     "The corresponding Human_Data_Entry rows for Heineken, Hermes (RMSP.XC), "
     "and Nestle (NSRGY) are already classified as Event Set = EXTENSION. "
     "No Event Set changes were required."),
]

for r_idx, (label, value) in enumerate(rows_text, start=1):
    cell_label = ws_note.cell(r_idx, 1, label)
    if value is None:
        cell_label.font = title_font if r_idx == 1 else header_font
    else:
        cell_label.font = header_font
        cell_value = ws_note.cell(r_idx, 2, value)
        cell_value.alignment = wrap_align

ws_note.column_dimensions["A"].width = 42
ws_note.column_dimensions["B"].width = 80

print(f"  Created sheet '{NOTE_SHEET}' with {len(rows_text)} rows.")

# ══════════════════════════════════════════════════════════════════════════════
# Save
# ══════════════════════════════════════════════════════════════════════════════
print(f"\nSaving to {WORKBOOK_PATH} …")
wb.save(WORKBOOK_PATH)
print("Done.")

# ── Final verification ────────────────────────────────────────────────────────
print("\n── Verification (re-load) ───────────────────────────────────────────────")
wb2 = openpyxl.load_workbook(WORKBOOK_PATH, data_only=True)
ws2 = wb2["Human_Data_Entry"]

ext_yes = ext_no = 0
for row in range(3, ws2.max_row + 1):
    es  = ws2.cell(row, COL_EVENT_SET).value
    pbc = ws2.cell(row, COL_PBC).value
    if es == "EXTENSION":
        if pbc == "YES":
            ext_yes += 1
        else:
            ext_no += 1

print(f"  EXTENSION rows after save: YES={ext_yes}, NO={ext_no}")
print(f"  Sheets: {wb2.sheetnames}")

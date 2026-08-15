"""Verify FROZEN no-match count and explain EXTENSION disagreements"""
import openpyxl
from datetime import date, datetime
from collections import defaultdict

WB_PATH = "/Users/nigelsim/Desktop/arp-master-4/data/workbook/Master_Data_CORRECTED_2026-08-14.xlsx"
wb = openpyxl.load_workbook(WB_PATH)
ws_h = wb['Human_Data_Entry']
ws_l = wb['LLM_Data_Entry']

def normalise_date(v):
    if v is None: return None
    if isinstance(v, (date, datetime)): return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try: return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except: pass
    return s

def to_float(v):
    if v is None: return None
    try: return float(v)
    except: return None

# Full LLM lookup (last row wins for duplicates)
llm_keys = {}
for rownum, row in enumerate(ws_l.iter_rows(min_row=3, values_only=True), start=3):
    ticker = str(row[1]).strip() if row[1] else ''
    docdate = normalise_date(row[9])
    if not ticker and not docdate: continue
    key = (ticker, docdate or '')
    llm_keys[key] = rownum

# Count Human rows by event_set
counts = defaultdict(lambda: defaultdict(int))
all_frozen = 0
for rownum, row in enumerate(ws_h.iter_rows(min_row=3, values_only=True), start=3):
    ticker = str(row[28]).strip() if row[28] else ''
    docdate = normalise_date(row[11])
    event_set = str(row[50]).strip() if row[50] else 'OTHER'
    if event_set not in ('FROZEN', 'EXTENSION'): event_set = 'OTHER'
    if not ticker and not docdate: continue
    key = (ticker, docdate or '')
    if key in llm_keys:
        counts[event_set]['match'] += 1
    else:
        counts[event_set]['no_match'] += 1
    if event_set == 'FROZEN':
        all_frozen += 1

print("Summary of Human rows by event set:")
for es in ('FROZEN', 'EXTENSION', 'OTHER'):
    c = counts[es]
    total = c['match'] + c['no_match']
    print(f"  {es}: total={total}, matched={c['match']}, no_match={c['no_match']}")
print(f"\nAll FROZEN Human rows: {all_frozen}")

# Show some EXTENSION that DO match
print("\n=== EXTENSION Human rows that match LLM (first 10) ===")
shown = 0
for rownum, row in enumerate(ws_h.iter_rows(min_row=3, values_only=True), start=3):
    ticker = str(row[28]).strip() if row[28] else ''
    docdate = normalise_date(row[11])
    event_set = str(row[50]).strip() if row[50] else ''
    if event_set != 'EXTENSION': continue
    key = (ticker, docdate or '')
    if key in llm_keys:
        company = row[0]
        h_pri = to_float(row[13])  # original N
        h_next = to_float(row[15]) # original P
        lrow = llm_keys[key]
        # get LLM prices
        lrow_data = list(ws_l.iter_rows(min_row=lrow, max_row=lrow, values_only=True))[0]
        l_pri = to_float(lrow_data[11])
        l_next = to_float(lrow_data[13])
        print(f"  row {rownum}: {ticker} {docdate} | {company}")
        print(f"    H_pri={h_pri}, L_pri={l_pri} | H_next={h_next}, L_next={l_next}")
        shown += 1
        if shown >= 5: break

# Show LLM EXTENSION rows for matching tickers (Colgate, Datadog)
print("\n=== LLM EXTENSION rows for CL/DDOG ===")
for rownum, row in enumerate(ws_l.iter_rows(min_row=3, values_only=True), start=3):
    ticker = str(row[1]).strip() if row[1] else ''
    event_set = str(row[21]).strip() if row[21] else ''
    if event_set == 'EXTENSION' and ticker in ('CL', 'DDOG', 'EBAY', 'SHOP', 'COST'):
        docdate = normalise_date(row[9])
        l_pri = to_float(row[11])
        l_next = to_float(row[13])
        print(f"  row {rownum}: {ticker} {docdate} | pri={l_pri}, next={l_next}")

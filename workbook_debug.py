"""Debug: why are all Human rows showing no-match?"""
import openpyxl
from datetime import date, datetime

WB_PATH = "/Users/nigelsim/Desktop/arp-master-4/data/workbook/Master_Data_CORRECTED_2026-08-14.xlsx"
wb = openpyxl.load_workbook(WB_PATH)

ws_h = wb['Human_Data_Entry']
ws_l = wb['LLM_Data_Entry']

# Print first 10 data rows of each sheet for key columns
print("=== Human_Data_Entry first 10 data rows (B,L,N,P,AY) ===")
for rownum, row in enumerate(ws_h.iter_rows(min_row=3, values_only=True), start=3):
    if rownum > 12:
        break
    ticker = row[1]   # B
    docdate = row[11]  # L (col 12, 0-indexed=11)
    priclose = row[13] # N
    nextopen = row[15] # P
    eventset = row[50] # AY
    print(f"  row{rownum}: ticker={ticker!r}, docdate={docdate!r} (type={type(docdate).__name__}), pri={priclose!r}, next={nextopen!r}, evset={eventset!r}")

print("\n=== LLM_Data_Entry first 10 data rows (B,J,L,N,V) ===")
for rownum, row in enumerate(ws_l.iter_rows(min_row=3, values_only=True), start=3):
    if rownum > 12:
        break
    ticker = row[1]   # B
    docdate = row[9]  # J (col 10, 0-indexed=9)
    priclose = row[11] # L
    nextopen = row[13] # N
    eventset = row[21] # V
    print(f"  row{rownum}: ticker={ticker!r}, docdate={docdate!r} (type={type(docdate).__name__}), pri={priclose!r}, next={nextopen!r}, evset={eventset!r}")

# Check what unique event set values are in Human
print("\n=== Human_Data_Entry unique EventSet values (col AY=51) ===")
eventsets = set()
for row in ws_h.iter_rows(min_row=3, values_only=True):
    v = row[50]
    eventsets.add(v)
print(f"  {sorted(str(v) for v in eventsets if v is not None)}")

print("\n=== LLM_Data_Entry unique EventSet values (col V=22) ===")
eventsets2 = set()
for row in ws_l.iter_rows(min_row=3, values_only=True):
    v = row[21]
    eventsets2.add(v)
print(f"  {sorted(str(v) for v in eventsets2 if v is not None)}")

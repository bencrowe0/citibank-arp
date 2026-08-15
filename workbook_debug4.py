"""Investigate: EXTENSION price discrepancies and no-match Human rows"""
import openpyxl
from datetime import date, datetime

WB_PATH = "/Users/nigelsim/Desktop/arp-master-4/data/workbook/Master_Data_CORRECTED_2026-08-14.xlsx"
wb = openpyxl.load_workbook(WB_PATH)
ws_h = wb['Human_Data_Entry']
ws_l = wb['LLM_Data_Entry']

def normalise_date(v):
    if v is None: return None
    if isinstance(v, (date, datetime)): return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try: return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except: pass
    return s

# Get all LLM event keys
llm_keys = set()
for row in ws_l.iter_rows(min_row=3, values_only=True):
    ticker = row[1]
    docdate = normalise_date(row[9])
    if ticker or docdate:
        llm_keys.add((str(ticker).strip() if ticker else '', docdate or ''))

# Get 25 no-match EXTENSION Human rows
print("=== 25 no-match EXTENSION Human rows (in Human but no LLM counterpart) ===")
count = 0
for rownum, row in enumerate(ws_h.iter_rows(min_row=3, values_only=True), start=3):
    ticker = row[28]  # AC = Re-priced listing
    docdate = normalise_date(row[11])  # L = Document Date
    event_set = str(row[50]).strip() if row[50] else ''
    if event_set != 'EXTENSION':
        continue
    key = (str(ticker).strip() if ticker else '', docdate or '')
    if key not in llm_keys:
        count += 1
        company = row[0]
        print(f"  Human row {rownum}: ticker={ticker!r}, date={docdate!r}, company={company!r}")
        if count >= 30:
            print("  (truncated)")
            break

print(f"\nTotal no-match EXTENSION rows: {count}")

# Investigate the 58 no-match FROZEN rows
print("\n=== 58 no-match FROZEN Human rows ===")
count2 = 0
for rownum, row in enumerate(ws_h.iter_rows(min_row=3, values_only=True), start=3):
    ticker = row[28]
    docdate = normalise_date(row[11])
    event_set = str(row[50]).strip() if row[50] else ''
    if event_set != 'FROZEN':
        continue
    key = (str(ticker).strip() if ticker else '', docdate or '')
    if key not in llm_keys:
        count2 += 1
        company = row[0]
        rater = row[6]  # col G = Rater
        if count2 <= 15:
            print(f"  Human row {rownum}: ticker={ticker!r}, date={docdate!r}, company={company!r}, rater={rater!r}")

print(f"\nTotal no-match FROZEN rows: {count2}")

# Check Lowe's FROZEN disagreement
print("\n=== Lowe's FROZEN re-priced price detail ===")
for rownum, row in enumerate(ws_h.iter_rows(min_row=3, values_only=True), start=3):
    ticker = row[28]
    if str(ticker).strip() == 'LOW':
        docdate = normalise_date(row[11])
        rep_pri = row[30]  # AE
        rep_next = row[32]  # AG
        event_set = row[50]
        print(f"  Human row {rownum}: date={docdate}, rep_pri={rep_pri}, rep_next={rep_next}, evset={event_set}")

for rownum, row in enumerate(ws_l.iter_rows(min_row=3, values_only=True), start=3):
    ticker = row[1]
    if str(ticker).strip() == 'LOW':
        docdate = normalise_date(row[9])
        rep_pri = row[25]  # Z
        rep_next = row[27]  # AB
        event_set = row[21]
        print(f"  LLM   row {rownum}: date={docdate}, rep_pri={rep_pri}, rep_next={rep_next}, evset={event_set}")

# Check what EXTENSION LLM rows look like for Colgate, Datadog etc
print("\n=== LLM EXTENSION rows — do they exist? ===")
llm_ext = []
for rownum, row in enumerate(ws_l.iter_rows(min_row=3, values_only=True), start=3):
    event_set = str(row[21]).strip() if row[21] else ''
    if event_set == 'EXTENSION':
        ticker = row[1]
        docdate = normalise_date(row[9])
        llm_ext.append((rownum, ticker, docdate))

print(f"LLM EXTENSION rows: {len(llm_ext)}")
for r in llm_ext[:15]:
    print(f"  row {r[0]}: {r[1]}, {r[2]}")

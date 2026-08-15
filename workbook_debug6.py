"""Final: understand Lowe's FROZEN disagree"""
import openpyxl
from datetime import date, datetime

WB_PATH = "/Users/nigelsim/Desktop/arp-master-4/data/workbook/Master_Data_CORRECTED_2026-08-14.xlsx"
wb = openpyxl.load_workbook(WB_PATH)
ws_h = wb['Human_Data_Entry']
ws_l = wb['LLM_Data_Entry']

def normalise_date(v):
    if v is None: return None
    if isinstance(v, (date, datetime)): return v.strftime("%Y-%m-%d")
    return str(v).strip()

print("=== All LLM rows for LOW ===")
for rownum, row in enumerate(ws_l.iter_rows(min_row=3, values_only=True), start=3):
    if str(row[1]).strip() == 'LOW':
        docdate = normalise_date(row[9])
        pri_orig = row[11]; next_orig = row[13]
        rep_pri = row[25]; rep_next = row[27]
        evset = row[21]
        print(f"  LLM row {rownum}: date={docdate}, orig_pri={pri_orig}, orig_next={next_orig}, rep_pri={rep_pri}, rep_next={rep_next}, evset={evset}")

print("\n=== All Human rows for LOW ===")
for rownum, row in enumerate(ws_h.iter_rows(min_row=3, values_only=True), start=3):
    if str(row[28]).strip() == 'LOW':
        docdate = normalise_date(row[11])
        pri_orig = row[13]; next_orig = row[15]
        rep_pri = row[30]; rep_next = row[32]
        evset = row[50]; rater = row[6]
        print(f"  Human row {rownum}: date={docdate}, orig_pri={pri_orig}, orig_next={next_orig}, rep_pri={rep_pri}, rep_next={rep_next}, evset={evset}, rater={rater}")

"""Debug: find where Ticker actually is in Human_Data_Entry"""
import openpyxl

WB_PATH = "/Users/nigelsim/Desktop/arp-master-4/data/workbook/Master_Data_CORRECTED_2026-08-14.xlsx"
wb = openpyxl.load_workbook(WB_PATH)
ws_h = wb['Human_Data_Entry']

# Print first 5 data rows, all columns, to find non-None values
print("=== Human_Data_Entry row 3 (first data row) — all columns ===")
row = list(ws_h.iter_rows(min_row=3, max_row=3, values_only=True))[0]
for col_idx, v in enumerate(row, start=1):
    if v is not None:
        ltr = openpyxl.utils.get_column_letter(col_idx)
        print(f"  col{col_idx}({ltr}): {v!r}")

print("\n=== Human_Data_Entry row 4 ===")
row = list(ws_h.iter_rows(min_row=4, max_row=4, values_only=True))[0]
for col_idx, v in enumerate(row, start=1):
    if v is not None:
        ltr = openpyxl.utils.get_column_letter(col_idx)
        print(f"  col{col_idx}({ltr}): {v!r}")

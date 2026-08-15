"""Debug: understand where ticker is in LLM sheet and what 'In LLM Universe' col looks like in Human"""
import openpyxl

WB_PATH = "/Users/nigelsim/Desktop/arp-master-4/data/workbook/Master_Data_CORRECTED_2026-08-14.xlsx"
wb = openpyxl.load_workbook(WB_PATH)

ws_l = wb['LLM_Data_Entry']
ws_h = wb['Human_Data_Entry']

# LLM row 3 — all non-None
print("=== LLM_Data_Entry row 3 — all non-None columns ===")
row = list(ws_l.iter_rows(min_row=3, max_row=3, values_only=True))[0]
for col_idx, v in enumerate(row, start=1):
    if v is not None:
        ltr = openpyxl.utils.get_column_letter(col_idx)
        print(f"  col{col_idx}({ltr}): {v!r}")

# Check Human_Data_Entry col AP (col 42) — "In LLM Universe"
# and col AC (col 29) — "Re-priced listing" (Ticker)
print("\n=== Human_Data_Entry: ticker (AC=29), docdate (L=12), 're-priced prior close date' (AD=30), 're-priced prior close' (AE=31), 're-priced next open' (AG=33), 'In LLM Universe' (AP=42) ===")
for rownum, row in enumerate(ws_h.iter_rows(min_row=3, values_only=True), start=3):
    if rownum > 10:
        break
    ticker_col29 = row[28]  # AC col29, 0-indexed=28
    docdate_l = row[11]     # L col12
    reprice_date = row[29]  # AD col30
    reprice_priclose = row[30]  # AE col31
    reprice_nextopen = row[32]  # AG col33
    in_llm = row[41]        # AP col42
    print(f"  row{rownum}: ticker={ticker_col29!r}, docdate={docdate_l!r}, reprice_date={reprice_date!r}, reprice_pri={reprice_priclose!r}, reprice_next={reprice_nextopen!r}, in_llm={in_llm!r}")

# Check LLM col Z (col 26) — 're-priced prior close'
print("\n=== LLM_Data_Entry: B=ticker, J=docdate, Y=reprice_close_date, Z=reprice_priclose, AB=reprice_nextopen, AK=in_human ===")
for rownum, row in enumerate(ws_l.iter_rows(min_row=3, values_only=True), start=3):
    if rownum > 10:
        break
    ticker = row[1]     # B
    docdate = row[9]    # J
    rep_close_date = row[24]  # Y col25
    rep_priclose = row[25]    # Z col26
    rep_nextopen = row[27]    # AB col28
    in_human = row[36]        # AK col37
    print(f"  row{rownum}: ticker={ticker!r}, docdate={docdate!r}, rep_close_date={rep_close_date!r}, rep_pri={rep_priclose!r}, rep_next={rep_nextopen!r}, in_human={in_human!r}")
